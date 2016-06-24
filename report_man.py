from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill, Fill
from openpyxl.styles import colors
import argparse



ft = Font(name='Arial', size=12)
ft_b = Font(name='Arial', size=12, bold=True)
headFill = PatternFill(start_color='FFA5C639',
                   fill_type='solid')

thin_boder = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


#['','',name,TestSuite,TestCase,appPackageName,abi]
def getTestInfo(failedItem):
	result = ['','',failedItem['name'],]
	for parent in failedItem.parents:
		if parent.name == 'TestCase':
			result.append(parent['name'])
		elif parent.name == 'TestSuite':
			result[3] = parent['name'] + '.' + result[3]
		elif parent.name == 'TestPackage':
			result.append(parent['appPackageName'])
			result.append(parent['abi'])
	return result

def getFirstTestInfo(failedItem):
	result = [failedItem['result'],failedItem['name'],]
	for parent in failedItem.parents:
		if parent.name == 'TestCase':
			result.append(parent['name'])
		elif parent.name == 'TestSuite':
			result[2] = parent['name'] + '.' + result[2]
		elif parent.name == 'TestPackage':
			result.append(parent['appPackageName'])
			result.append(parent['abi'])
	return result


def findTestFail(soup):
	failedTest = soup.find_all('Test', result='fail')
	result = [('Status','PIC', 'Test', 'Test Case', 'Test Package', 'abi'),]

	for item in failedTest:
		result.append(getTestInfo(item))
	return result

def findTestNotPass(soup):
	failedTest = soup.find_all('Test', result='fail')
	notExeTest = soup.find_all('Test', result='notExecuted')
	failedTest.extend(notExeTest)
	result = [('Result','Test', 'Test Case', 'Test Package', 'abi'),]

	for item in failedTest:
		result.append(getFirstTestInfo(item))
	return result


def writeTestFail(ws, failedItem):
	ws['A13'] = 'Failed Test Cases'
	ws['A13'].font = ft_b
	
	for i, item in enumerate(failedItem, start=15):
		for j, col in enumerate(item, start=1):
			c = ws.cell(row=i, column=j)
			c.value = item[len(item)-j]
			if i == 15:
				applyHeadStyle(c)
			else:
				applyBorder(c)

	return ws


#def detailedBlock(soup):

def applyHeadStyle(cell):
	cell.font = ft_b
	cell.fill = headFill
	cell.border = thin_boder

def applyBorder(cell):
	cell.font = ft
	cell.border = thin_boder


def writeDetailed(ws, soup):
	failedTests = soup.find_all('Test', result='fail')
	result = {}
	rowi = 0

	for failedTest in failedTests:
		testInfo = getTestInfo(failedTest)
		packageKey = 'Compatibility Test Package: ' + testInfo[4] + ' ABI: ' + testInfo[5]
		if packageKey in result.keys():
			if testInfo[3] in result[packageKey].keys():
				result[packageKey][testInfo[3]].append([testInfo[2], failedTest.FailedScene['message']])
			else:
				result[packageKey][testInfo[3]] = [[testInfo[2], failedTest.FailedScene['message']],]
		else:
			result[packageKey] = {testInfo[3]:[[testInfo[2], failedTest.FailedScene['message']],]}


	for packageKey in result.keys():

		rowi +=1
		ws.append([])
		#print packageKey
		#ws.append([packageKey,])
		rowi +=1
		ws.cell(row=rowi, column=1).value = packageKey
		ws.cell(row=rowi, column=1).font = ft_b
		ws.merge_cells(start_row=rowi,start_column=1,end_row=rowi,end_column=2)


		rowi +=1
		ws.append(['Test', 'Details'])
		c = ws.cell(row=rowi, column=1)
		applyHeadStyle(c)

		c = ws.cell(row=rowi, column=2)
		applyHeadStyle(c)

		for testsuit in result[packageKey].keys():

			#ws.append([testsuit,])
			rowi +=1
			ws.merge_cells(start_row=rowi,start_column=1,end_row=rowi,end_column=2)
			c = ws.cell(row=rowi, column=1)
			c.value = testsuit
			applyBorder(c)
			c = ws.cell(row=rowi, column=2)
			applyBorder(c)

			for test in result[packageKey][testsuit]:
				rowi +=1
				c = ws.cell(row=rowi, column=1)
				c.value = test[0]
				applyBorder(c)
				c = ws.cell(row=rowi, column=2)
				c.value = test[1]
				applyBorder(c)
				#ws.append(test)

		rowi +=1
		ws.append([])

	return ws


def writeTestSummary(ws, values):
	colName = ('CTS version','Test timeout','Host Info',
		'Plan name','Start time','End time','Tests Passed',
		'Tests Failed','Tests Timed out','Tests Not Executed')

	ws.merge_cells('A1:B1')
	ws['A1'] = "Test Summary"
	applyHeadStyle(ws['A1'])
	applyHeadStyle(ws['B1'])

	for h, v in zip(colName, values):
		ws.append([h,v])

	return ws



def findTestSummary(soup):
	#version,timeout,info,name,start,end,passed,failed,timeout,notexe
	result = (soup.TestResult.HostInfo.Cts["version"],
		soup.TestResult.HostInfo.Cts.IntValue["value"],
		soup.TestResult.HostInfo["name"] + "(" + 
			soup.TestResult.HostInfo.Os["name"] + "-" + 
			soup.TestResult.HostInfo.Os["version"] + ")",
		soup.TestResult["testPlan"],
		soup.TestResult["starttime"],
		soup.TestResult["endtime"],
		soup.TestResult.Summary["pass"],
		soup.TestResult.Summary["failed"],
		soup.TestResult.Summary["timeout"],
		soup.TestResult.Summary["notExecuted"]
		)

	return result

def resizeCol(ws):
	colWidth = 10
	for i, col in enumerate(ws.columns,start=1):
		for j in range(1, len(col)):
			foo = ws.cell(row=j, column=i).value
			if foo is None:
				continue
			if colWidth < len(foo):
				colWidth = len(foo)

		ws.column_dimensions[get_column_letter(i)].width = colWidth
		colWidth = 10
	return ws



def main():
	parser = argparse.ArgumentParser()
	parser.add_argument("-f", dest="first_name", required=True) 
	parser.add_argument("-l", dest="last_name", required=True)
	parser.add_argument("-o", dest="output", default="output.xlsx")

	args = parser.parse_args()

	first_soup = bs(open(args.first_name),'xml')

	#print soup

	wb = Workbook()

	ws = wb.active
	ws.title = 'First Run'
	summaryInfo = findTestSummary(first_soup)
	ws = writeTestSummary(ws, summaryInfo)

	failedItem = findTestNotPass(first_soup)
	ws = writeTestFail(ws, failedItem)

	ws = resizeCol(ws)


	last_soup = bs(open(args.last_name),'xml')
	ws = wb.create_sheet()
	ws.title = 'Last Run'
	summaryInfo = findTestSummary(last_soup)
	ws = writeTestSummary(ws, summaryInfo)
	failedItem = findTestFail(last_soup)
	ws = writeTestFail(ws, failedItem)
	ws = resizeCol(ws)


	ws_detail = wb.create_sheet()
	ws_detail.title = 'Detailed Test Report'

	ws_detail = writeDetailed(ws_detail, last_soup)
	ws_detail = resizeCol(ws_detail)


	wb.save(args.output)


if __name__ == "__main__":
	main()